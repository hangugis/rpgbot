import random
import discord
import openpyxl
# noinspection PyUnresolvedReferences
import time

client = discord.Client()

@client.event
async def on_ready():
    print(client.user.id)
    print("ready")
    game = discord.Game("슬롯머신 개장 문의 한구깃#1098")
    await client.change_presence(status=discord.Status.online, activity=game)

@client.event
async def on_message(message):
    global hand1
    global hand2
    global dsum
    global sum
    global plus
    global teest
    if message.content.startswith("테스트"):
        await message.channel.send("확인")
    if message.content.startswith("뭘 넣지"):
        await message.channel.send("이스터에그 발견!")
    if message.content == "조금 느린듯":
        await message.channel.send("바위\n제가 이겼습니다")
    if message.content == "$도움말":
        embed = discord.Embed(title="도움말", description="제작중...", color=0x62c1cc)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
        embed.add_field(name="가위바위보", value="`$가위`**,**`$바위`**,**`$보`", inline=False)
        embed.add_field(name="대장간", value="`$대장간가입`**,**`$강화`**,**`$상태`**", inline=False)
        embed.add_field(name="rpg", value="`$참가`**,**`$인벤토리`**,**`$던전`**,**`추가예정`", inline=False)
        embed.set_footer(text="문의는 한구깃#")  # 하단에 들어가는 조그마한 설명을 잡아줍니다
        await message.channel.send(embed=embed)  # embed를 포함 한 채로 메시지를 전송합니다.
        # await message.channel.send("할 말", embed=embed)  # embed와 메시지를 함께 보내고 싶으시면 이렇게 사용하시면 됩니다.
    if message.content == "$test":
        embed = discord.Embed(title="테스트용", description="https://www.youtube.com/watch?v=4xl9F2zEx34", color=0x62c1cc)
        embed.set_footer(text="")
        await message.channel.send(embed=embed)
    if message.content == "$가위" or message.content == "$바위" or message.content == "$보":
        bot_response = random.randint(1, 3)
        if bot_response == 1:
            if message.content == "$가위":
                await message.channel.send("가위")
                await message.channel.send("비겼습니다")
            elif message.content == "$바위":
                await message.channel.send("가위")
                await message.channel.send("제가 졌습니다")
            else:
                await message.channel.send("가위")
                await message.channel.send("제가 이겼습니다")
        elif bot_response == 2:
            if message.content == "$가위":
                await message.channel.send("바위")
                await message.channel.send("제가 이겼습니다")
            elif message.content == "$바위":
                await message.channel.send("바위")
                await message.channel.send("비겼습니다")
            else:
                await message.channel.send("바위")
                await message.channel.send("제가 졌습니다")
        else:
            if message.content == "$가위":
                await message.channel.send("보")
                await message.channel.send("제가 졌습니다")
            elif message.content == "$바위":
                await message.channel.send("보")
                await message.channel.send("제가 이겼습니다")
            else:
                await message.channel.send("보")
                await message.channel.send("비겼습니다")
    if message.content.startswith("$분석"):
        count = 0
        l = 0
        f = 0
        while l <100000:
            a = random.randint(1,100)
            b = random.randint(1,100)
            c = random.randint(1,100)
            sum  = a+b+c
            if sum >= 240:
                count += 1
            l += 1
        f = count/1000
        await message.channel.send(str(f)+"%")

    if message.content == "$슬롯머신":
        one = random.randint(1,9)
        two = random.randint(1,9)
        three = random.randint(1,9)
        if one == 7:
            one1 = ":o:"
        else:
            one1 = ":x:"
        if two == 7:
            two1 = ":o:"
        else:
            two1 = ":x:"
        if three == 7:
            three1 = ":o:"
        else:
            three1 = ":x:"
        embed = discord.Embed(title="슬롯머신", description="1회당 100만메소", color=0x62c1cc)
        embed.add_field(name=one1, value=one,inline=True)
        embed.add_field(name=two1,value=two, inline=True)
        embed.add_field(name=three1,value=three, inline=True)
        embed.set_footer(text="당첨시 3억메소")
        await message.channel.send(embed=embed)
    if message.content == "$슬롯머신 시뮬":
        one2 = random.randint(1,9)
        two2 = random.randint(1,9)
        three2 = random.randint(1,9)
        if one2 == 7:
            one3 = ":o:"
        else:
            one3 = ":x:"
        if two2 == 7:
            two3 = ":o:"
        else:
            two3 = ":x:"
        if three2 == 7:
            three3 = ":o:"
        else:
            three3 = ":x:"
        embed = discord.Embed(title="슬롯머신 시뮬", description="무료사용가능", color=0x62c1cc)
        embed.add_field(name=one3, value=one2,inline=True)
        embed.add_field(name=two3,value=two2, inline=True)
        embed.add_field(name=three3,value=three2, inline=True)
        embed.set_footer(text="(시뮬레이션)")
        await message.channel.send(embed=embed)
    if message.content == "$블랙잭 규칙":
        await message.channel.send("블랙잭 규칙 \n카드 두 장을 기본적으로 지급받게 되며,\n카드 숫자를 합쳐 가능한 21에 가깝게 만들면 이기는 게임이다.\n처음 받은 2장 합쳐 21이 나오는 경우 블랙잭이 되며 승리하게 된다.")
        await message.channel.send("21이 되지 않았을 경우 원한다면 얼마든지 카드를 계속 뽑을 수 있다\n하지만 카드 숫자의 합이 21을 초과하게 되는 순간 '버스트' 라고 하며 딜러의 결과에 관계없이 플레이어가 패배한다.")
        await message.channel.send("K,Q,J 카드는 10으로 처리가 되며 A는 11로 처리가 됩니다.")
    if message.content == "$테스트":
        teest = 0
        await message.channel.send(str(teest))
        teest=3
    if message.content == "$테스트2":
        await message.channel.send(str(teest))
    if message.content == "$블랙잭":
        card1 = random.randint(1, 11)
        card2 = random.randint(1, 11)
        hand1 = 0
        hand2 = 0
        if card1 < 11:
            hand1 = card1
        if card2 < 11:
            hand2 = card2
        if card1 == 11:
            hand1 = "Ace"
        if card1 == 10:
            card10 = random.randint(1, 4)
            if card10 == 1:
                hand1 = "10"
            if card10 == 2:
                hand1 = "King"
            if card10 == 3:
                hand1 = "Queen"
            if card10 == 4:
                hand1 = "Jack"
        if card2 == 11:
            hand2 = "Ace"
        if card2 == 10:
            card10 = random.randint(1, 4)
            if card10 == 1:
                hand2 = "10"
            if card10 == 2:
                hand2 = "King"
            if card10 == 3:
                hand2 = "Queen"
            if card10 == 4:
                hand2 = "Jack"
        sum = card1 + card2
        await message.channel.send("플레이어의 패는" + str(hand1) + "," + str(hand2) + "입니다.\n합은" + str(sum) + "입니다.")
        dcard1 = random.randint(1, 11)
        dcard2 = random.randint(1, 11)
        dhand1 = 0
        dhand2 = 0
        if dcard1 < 11:
            dhand1 = dcard1
        if dcard2 < 11:
            dhand2 = dcard2
        if dcard1 == 11:
            dhand1 = "Ace"
        if dcard1 == 10:
            dcard10 = random.randint(1, 4)
            if dcard10 == 1:
                dhand1 = "10"
            if dcard10 == 2:
                dhand1 = "King"
            if dcard10 == 3:
                dhand1 = "Queen"
            if dcard10 == 4:
                dhand1 = "Jack"
        if dcard2 == 11:
            dhand2 = "Ace"
        if dcard2 == 10:
            dcard10 = random.randint(1, 4)
            if dcard10 == 1:
                dhand2 = "10"
            if dcard10 == 2:
                dhand2 = "King"
            if dcard10 == 3:
                dhand2 = "Queen"
            if dcard10 == 4:
                dhand2 = "Jack"
        dsum = dcard1 + dcard2
        await message.channel.send("딜러의 패는" + str(dhand1) + "입니다.")
        if sum == 21:
            if dsum == 21:
                await message.channel.send("딜러와 플레이어가 둘 다 블랙잭입니다.\n(무승부)")
            else:
                await message.channel.send("플레이어가 블랙잭으로 승리하셨습니다.\n(승리x1.5)")
        elif dsum == 21:
            await message.channel.send("딜러가 블랙잭이므로 패배입니다.\n(패배)")
    if message.content == "$STAY":
        if dsum < 17:
            plus = random.randint(1, 11)
            dsum = dsum + plus
            if dsum > 21:
                await message.channel.send("딜러의 합이 21이 넘었으므로 플레이어의 승리입니다.\n(승리x2)")
            elif dsum == 21:
                await message.channel.send("딜러의 합이 21이므로 딜러의 승리입니다.\n(패배)")
            elif dsum < sum:
                await message.channel.send(
                    "딜러의 합은 " + str(dsum) + ", 플레이어의 합은 " + str(sum) + "이므로 플레이어의 승리입니다\n(승리x2)")
            elif sum < dsum:
                await message.channel.send(
                    "딜러의 합은 " + str(dsum) + ", 플레이어의 합은 " + str(sum) + "이므로 딜러의 승리입니다\n(패배)")
        elif dsum > 16:
            if dsum > 21:
                await message.channel.send("딜러의 합이 21이 넘었으므로 플레이어의 승리입니다.\n(승리x2)")
            elif dsum == 21:
                await message.channel.send("딜러의 합이 21이므로 딜러의 승리입니다.\n(패배)")
            elif dsum < sum:
                await message.channel.send(
                    "딜러의 합은 " + str(dsum) + ", 플레이어의 합은 " + str(sum) + "이므로 플레이어의 승리입니다\n(승리x2)")
            elif sum < dsum:
                await message.channel.send(
                    "딜러의 합은 " + str(dsum) + ", 플레이어의 합은 " + str(sum) + "이므로 딜러의 승리입니다\n(패배)")
    if message.content == "$HIT":
        plus1 = random.randint(1,14)
        sum = sum + plus1
        if plus1 > 11:
            plus10 = random.randint(2, 4)
            if plus10 == 2:
                plus1 = "King"
            if plus10 == 3:
                plus1 = "Queen"
            if plus10 == 4:
                plus1 = "Jack"
        if dsum < 17:
            plus2 = random.randint(1, 11)
            dsum = dsum + plus2
            if sum > 21:
                if dsum < 21:
                    await message.channel.send(str(plus1) + "를 더 받았습니다. 합은" + str(sum))
                    await message.channel.send("플레이어의 합이 21이 넘었으므로 버스트입니다.\n(패배)")
                elif dsum > 21:
                    await message.channel.send(str(plus1) + "를 더 받았습니다. 합은" + str(sum))
                    await message.channel.send("플레이어의 합이 21이 넘었으므로 버스트입니다.\n(패배)")
            elif sum == 21:
                if dsum == 21:
                    await message.channel.send("플레이어와 딜러의 합이 모두 21이므로 무승부입니다.\n(무승부)")
                else:
                    await message.channel.send("플레이어의 합이 21이므로 플레이어의 승리입니다\n(승리x2)")
            elif sum < 21:
                await message.channel.send(str(plus1)+"를 더 받았습니다. 합은"+str(sum))
        elif dsum > 16:
            if sum > 21:
                if dsum < 21:
                    await message.channel.send(str(plus1) + "를 더 받았습니다. 합은" + str(sum))
                    await message.channel.send("플레이어의 합이 21이 넘었으므로 버스트입니다.\n(패배)")
                elif dsum > 21:
                    await message.channel.send(str(plus1) + "를 더 받았습니다. 합은" + str(sum))
                    await message.channel.send("플레이어의 합이 21이 넘었으므로 버스트입니다.\n(무승부)")
            elif sum == 21:
                if dsum == 21:
                    await message.channel.send("플레이어와 딜러의 합이 모두 21이므로 무승부입니다.\n(무승부)")
                else:
                    await message.channel.send("플레이어의 합이 21이므로 플레이어의 승리입니다\n(승리x2)")
            elif sum < 21:
                await message.channel.send(str(plus1)+"를 더 받았습니다. 합은"+str(sum))
    if message.content.startswith("$대장간가입"):
        file = openpyxl.load_workbook("강화.xlsx")
        sheet = file.active
        a = 0
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-":
               sheet["A" + str(i)].value = str(message.author)
               sheet["B"+str(i)].value = 1
               a = 2
               break
            elif sheet["A"+str(i)].value == str(message.author):
                a = 1
                break
        file.save("강화.xlsx")
        if a == 1:
            await message.channel.send("이미 가입되셨습니다.")
        elif a == 2:
            await message.channel.send("가입이 완료되었습니다.")
    if message.content.startswith("$상태"):
        b=0
        a=0
        file = openpyxl.load_workbook("강화.xlsx")
        sheet = file.active
        for i in range(1,51):
            if sheet["A"+str(i)].value == str(message.author):
                b=1
                a = sheet["B"+str(i)].value
                break
            else:
                b=2
        file.save("강화.xlsx")
        if b == 1:
            await message.channel.send(str(a)+"성입니다.")
        elif b == 2:
            await message.channel.send("대장간에 가입을 해주세요!")
    if message.content == "$강화":
        file = openpyxl.load_workbook("강화.xlsx")
        sheet = file.active
        a = 0
        b = 0
        c = str(message.author)
        d = c[:-5]
        p = random.randint(1, 100)
        for i in range(1,51):
            if sheet["A"+str(i)].value == str(message.author) or str(c) == sheet["A"+str(i)]:
                if sheet["C"+str(i)].value < 2:
                    if sheet["B"+str(i)].value < 15:
                        if p < 95 -5*sheet["B"+str(i)].value:
                            sheet["B" + str(i)].value = sheet["B" + str(i)].value + 1
                            a = sheet["B" + str(i)].value
                            b = 1
                            sheet["C" + str(i)].value = 0
                            break
                        else:
                            if sheet["B"+str(i)].value < 11:
                                a = sheet["B" + str(i)].value
                                b=3
                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1
                                break
                            elif 5 < sheet["B"+str(i)].value and sheet["B"+str(i)].value != 10:
                                sheet["B" + str(i)].value = sheet["B" + str(i)].value - 1
                                a = sheet["B" + str(i)].value
                                b=3
                                sheet["C" + str(i)].value = sheet["C"+str(i)].value + 1
                                break
                    if 14 < sheet["B"+str(i)].value < 22:
                        if p < 31:
                            sheet["B" + str(i)].value = sheet["B" + str(i)].value + 1
                            a = sheet["B" + str(i)].value
                            b = 1
                            sheet["C" + str(i)].value = 0
                            break
                        elif 30 < p < 36:
                            sheet["B" + str(i)].value = 0
                            a = sheet["B" + str(i)].value
                            b = 4
                            break
                        else:
                            if sheet["B"+str(i)].value == 20 or sheet["B"+str(i)].value == 15:
                                a = sheet["B" + str(i)].value
                                b=3
                                break
                            else:
                                sheet["B" + str(i)].value = sheet["B" + str(i)].value - 1
                                a = sheet["B" + str(i)].value
                                b=3
                                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1
                                break
                    if sheet["B"+str(i)].value == 22:
                        if p < 4:
                            sheet["B" + str(i)].value = sheet["B" + str(i)].value + 1
                            a = sheet["B" + str(i)].value
                            b = 1
                            await message.channel.send(str(d) + "님 23성 축하드립니다!이 말을 스샷을 찍어 상금 23억을 받아가세요")
                            break
                        elif 3 < p < 24:
                            sheet["B" + str(i)].value = 0
                            a = sheet["B" + str(i)].value
                            b = 4
                            break
                        else:
                            sheet["B" + str(i)].value = sheet["B" + str(i)].value - 1
                            a = sheet["B" + str(i)].value
                            b = 3
                            break
                elif sheet["C"+str(i)].value == 2:
                    sheet["B" + str(i)].value = sheet["B" + str(i)].value + 1
                    a = sheet["B" + str(i)].value
                    b = 5
                    sheet["C" + str(i)].value = 0
                    break
            else:
                b = 2
        file.save("강화.xlsx")
        if b == 1:
            await message.channel.send("강화에 성공했습니다!\n"+str(a)+"성입니다")
        elif b == 2:
            await message.channel.send("대장간에 가입을 해주세요!")
        elif b == 3:
            await message.channel.send("강화에 실패했습니다!\n"+str(a)+"성입니다")
        elif b == 4:
            await message.channel.send("무기가 파괴되었습니다!")
        elif b == 5:
            await message.channel.send("찬스타임!\n"+str(a)+"성입니다")

    if message.content.startswith("$참가"):
        file=openpyxl.load_workbook("인벤토리.xlsx")
        sheet = file.active
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-":
               sheet["A" + str(i)].value = str(message.author)
               a = 2
               break
            elif sheet["A"+str(i)].value == str(message.author):
                a = 1
                break
        file.save("인벤토리.xlsx")
        if a == 1:
            await message.channel.send("이미 가입되셨습니다.")
        elif a == 2:
            await message.channel.send("가입이 완료되었습니다.")
        file=openpyxl.load_workbook("경험치.xlsx")
        sheet = file.active
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-":
               sheet["A" + str(i)].value = str(message.author)
               a = 2
               break
            elif sheet["A"+str(i)].value == str(message.author):
                a = 1
                break
        file.save("경험치.xlsx")
    if message.content.startswith("$인벤토리"):
        file = openpyxl.load_workbook("인벤토리.xlsx")
        sheet = file.active
        a = str(message.author)
        b = a[:-5]
        for i in range(1, 51):
            if sheet["A"+str(i)].value == str(message.author):
                if sheet["B"+str(i)].value > 0:
                    await message.channel.send(str(b) + "님의 인벤토리")
                    await message.channel.send(str(sheet["B"+str(51)].value)+":"+str(sheet["B"+str(i)].value)+"개")
                if sheet["C"+str(i)].value > 0:
                    await message.channel.send(str(sheet["C"+str(51)].value)+":"+str(sheet["C"+str(i)].value)+"개")
                if sheet["D"+str(i)].value > 0:
                    await message.channel.send(str(sheet["D"+str(51)].value)+":"+str(sheet["D"+str(i)].value)+"개")
                if sheet["E" + str(i)].value > 0:
                    await message.channel.send(str(sheet["E" + str(51)].value) + ":" + str(sheet["E" + str(i)].value) + "개")
                if sheet["F" + str(i)].value > 0:
                    await message.channel.send(str(sheet["F" + str(51)].value) + ":" + str(sheet["F" + str(i)].value) + "개")
                if sheet["G" + str(i)].value > 0:
                    await message.channel.send(str(sheet["G" + str(51)].value) + ":" + str(sheet["G" + str(i)].value) + "개")
                if sheet["H" + str(i)].value > 0:
                    await message.channel.send(str(sheet["H" + str(51)].value) + ":" + str(sheet["H" + str(i)].value) + "개")
                if sheet["I" + str(i)].value > 0:
                    await message.channel.send(str(sheet["I" + str(51)].value) + ":" + str(sheet["I" + str(i)].value) + "개")
                if sheet["J" + str(i)].value > 0:
                    await message.channel.send(str(sheet["J" + str(51)].value) + ":" + str(sheet["J" + str(i)].value) + "개")
                if sheet["K" + str(i)].value > 0:
                    await message.channel.send(str(sheet["K" + str(51)].value) + ":" + str(sheet["K" + str(i)].value) + "개")
                if sheet["L" + str(i)].value > 0:
                    await message.channel.send(str(sheet["L" + str(51)].value) + ":" + str(sheet["L" + str(i)].value) + "개")
                if sheet["M" + str(i)].value > 0:
                    await message.channel.send(str(sheet["M" + str(51)].value) + ":" + str(sheet["M" + str(i)].value) + "개")
                if sheet["N" + str(i)].value > 0:
                    await message.channel.send(str(sheet["N" + str(51)].value) + ":" + str(sheet["N" + str(i)].value) + "개")
                if sheet["O" + str(i)].value > 0:
                    await message.channel.send(str(sheet["O" + str(51)].value) + ":" + str(sheet["O" + str(i)].value) + "개")
                if sheet["P" + str(i)].value > 0:
                    await message.channel.send(str(sheet["P" + str(51)].value) + ":" + str(sheet["P" + str(i)].value) + "개")
                if sheet["Q" + str(i)].value > 0:
                    await message.channel.send(str(sheet["Q" + str(51)].value) + ":" + str(sheet["Q" + str(i)].value) + "개")
                if sheet["R" + str(i)].value > 0:
                    await message.channel.send(str(sheet["R" + str(51)].value) + ":" + str(sheet["R" + str(i)].value) + "개")
                if sheet["S" + str(i)].value > 0:
                    await message.channel.send(str(sheet["S" + str(51)].value) + ":" + str(sheet["S" + str(i)].value) + "개")
                if sheet["T" + str(i)].value > 0:
                    await message.channel.send(str(sheet["T" + str(51)].value) + ":" + str(sheet["T" + str(i)].value) + "개")
                if sheet["U" + str(i)].value > 0:
                    await message.channel.send(str(sheet["U" + str(51)].value) + ":" + str(sheet["U" + str(i)].value) + "개")
                if sheet["V" + str(i)].value > 0:
                    await message.channel.send(str(sheet["V" + str(51)].value) + ":" + str(sheet["V" + str(i)].value) + "개")
                    break
            elif sheet["A"+str(i)].value != str(message.author):
                await message.channel.send(str(b)+"님의 인벤토리는 비어있습니다.")
                break
        file.save("인벤토리.xlsx")
    if message.content.startswith("$던전"):
        await message.channel.send("추가 예정입니다.")






















client.run("NzgxODk2NzYzMjQxNzkxNTQw.X8EUGw.kniENR7788zwRU8QTELoqQlPosE")
